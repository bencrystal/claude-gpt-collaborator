from openai import AsyncOpenAI
import base64
import io
import json
import os

client = AsyncOpenAI(api_key=os.environ.get("OPENAI_API_KEY"))

MODEL = "gpt-4o"


_EXCEL_TYPES = {
    "application/vnd.ms-excel",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}


def _excel_to_text(data_b64: str, name: str) -> str:
    """Convert a base64-encoded Excel file to a CSV-like text representation."""
    import openpyxl
    raw = base64.b64decode(data_b64)
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(",".join("" if v is None else str(v) for v in row))
        parts.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows))
    return "\n\n".join(parts)


def _build_content_parts(prompt: str, files: list[dict]) -> list:
    """Build OpenAI message content parts from prompt + file attachments."""
    parts = []
    text_file_parts = []

    for f in files:
        mt = f["media_type"]
        if mt.startswith("image/"):
            parts.append({
                "type": "image_url",
                "image_url": {"url": f"data:{mt};base64,{f['data']}"},
            })
        elif mt in _EXCEL_TYPES or f["name"].lower().endswith((".xls", ".xlsx")):
            try:
                text = _excel_to_text(f["data"], f["name"])
            except Exception as e:
                text = f"[Could not parse Excel file: {e}]"
            text_file_parts.append(f"=== {f['name']} ===\n{text}")
        else:
            try:
                text = base64.b64decode(f["data"]).decode("utf-8", errors="replace")
            except Exception:
                text = f["data"]
            text_file_parts.append(f"=== {f['name']} ===\n{text}")

    full_text = prompt
    if text_file_parts:
        full_text = "\n\n".join(text_file_parts) + "\n\n" + prompt

    parts.append({"type": "text", "text": full_text})
    return parts


async def generate_initial(prompt: str, files: list[dict] = None) -> str:
    content = _build_content_parts(prompt, files or [])
    response = await client.chat.completions.create(
        model=MODEL,
        max_tokens=4096,
        messages=[
            {
                "role": "system",
                "content": (
                    "You are a thorough, expert assistant. "
                    "Respond to the prompt as completely and insightfully as possible."
                ),
            },
            {"role": "user", "content": content},
        ],
    )
    return response.choices[0].message.content


async def generate_critique_and_improve(
    original_prompt: str,
    your_previous: str,
    other_previous: str,
    other_name: str,
) -> dict:
    """Returns {"score": int, "missing": [str], "response": str}"""
    user_message = f"""Original prompt:
{original_prompt}

Your previous response:
{your_previous}

{other_name}'s response to the same prompt:
{other_previous}

Instructions:
1. Score {other_name}'s response out of 10 for how completely it answers the original prompt.
2. List 3-5 specific things that were absent or underdeveloped in {other_name}'s response.
3. Produce an improved response that is more complete than either previous response.

Return a JSON object with exactly these keys:
- "score": integer 1-10
- "missing": array of 3-5 short strings
- "response": your full improved response as a string"""

    response = await client.chat.completions.create(
        model=MODEL,
        max_tokens=4096,
        response_format={"type": "json_object"},
        messages=[
            {
                "role": "system",
                "content": (
                    "You are a thorough, expert assistant engaged in a collaborative refinement process. "
                    "Always respond with valid JSON containing 'score', 'missing', and 'response' keys."
                ),
            },
            {"role": "user", "content": user_message},
        ],
    )
    return json.loads(response.choices[0].message.content)
